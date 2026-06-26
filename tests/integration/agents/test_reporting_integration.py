"""Integration tests for the Reporting Agent handler.

Tests the full message-to-completion flow using realistic mocks for all I/O:
  - DB repositories (AsyncMock, mimicking real DB query results)
  - Azure Blob Storage uploader
  - Key Vault client
  - Webhook HTTP delivery
  - Excel / PDF generators (real calls — tests actual bytes output)

No live Azure services or PostgreSQL are required.
"""

from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from waf_reporting.aggregator import AggregatedReport, FindingAggregator, PillarSummary
from waf_reporting.excel_generator import ExcelGenerator
from waf_reporting.handler import ReportingHandler
from waf_reporting.pdf_generator import PdfGenerator
from waf_reporting.storage_uploader import StorageUploader
from waf_reporting.webhook_service import WebhookDeliveryError, WebhookService
from waf_shared.domain.events.assessment_events import ReportingRequestedEvent
from waf_shared.domain.events.base import CloudEventEnvelope
from waf_shared.domain.models.assessment import Assessment, AssessmentStatus
from waf_shared.domain.models.finding import Finding, FindingStatus, Severity
from waf_shared.domain.models.report import AssessmentReport, AssessmentSummary
from waf_shared.domain.models.webhook import TenantWebhookEndpoint

# ── Constants ─────────────────────────────────────────────────────────────────

_TENANT_ID = uuid.UUID("dddd0000-0000-0000-0000-000000000001")
_ASSESSMENT_ID = uuid.UUID("eeee0000-0000-0000-0000-000000000002")
_BATCH_ID = uuid.UUID("ffff0000-0000-0000-0000-000000000003")
_REPORT_ID = uuid.UUID("a1a10000-0000-0000-0000-000000000004")

_XLSX_PATH = f"reports/{_TENANT_ID}/{_ASSESSMENT_ID}/report.xlsx"
_PDF_PATH = f"reports/{_TENANT_ID}/{_ASSESSMENT_ID}/report.pdf"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _make_assessment(status: AssessmentStatus = AssessmentStatus.REPORTING) -> Assessment:
    return Assessment(
        id=_ASSESSMENT_ID,
        tenant_id=_TENANT_ID,
        idempotency_key="key-int-test",
        status=status,
        subscription_ids=["sub-int-1"],
        pillar_filter=None,
        tag_filter=None,
        requested_by_oid="oid-int",
        total_batches=2,
        completed_batches=2,
        cancellation_requested_at=None,
        created_at=datetime.now(UTC),
        updated_at=datetime.now(UTC),
    )


def _make_findings(count: int = 5) -> list[Finding]:
    severities = ["critical", "high", "medium", "low", "informational"]
    return [
        Finding(
            id=uuid.uuid4(),
            assessment_id=_ASSESSMENT_ID,
            batch_id=_BATCH_ID,
            tenant_id=_TENANT_ID,
            rule_id=f"WAF-SEC-{i:03d}",
            resource_id=f"/subs/s/rgs/rg/providers/Resource/{i}",
            resource_type="Microsoft.Network/applicationGateways",
            status=FindingStatus.OPEN,
            severity=Severity(severities[i % len(severities)]),
            pillar="security" if i % 2 == 0 else "reliability",
            confidence_score=0.85,
            title=f"Finding {i}",
            recommendation=f"Recommendation {i}",
            evidence={"result": "FAIL"},
            evaluation_type="deterministic",
            created_at=datetime.now(UTC),
        )
        for i in range(count)
    ]


def _make_aggregated(findings: list[Finding]) -> AggregatedReport:
    by_sev: dict[str, int] = {}
    by_pillar: dict[str, dict[str, int]] = {}
    for f in findings:
        by_sev[f.severity.value] = by_sev.get(f.severity.value, 0) + 1
        if f.pillar not in by_pillar:
            by_pillar[f.pillar] = {}
        by_pillar[f.pillar][f.severity.value] = (
            by_pillar[f.pillar].get(f.severity.value, 0) + 1
        )
    pillars = {
        p: PillarSummary(
            pillar=p,
            findings_by_severity=sev_map,
            total_findings=sum(sev_map.values()),
            compliance_score=0.6,
        )
        for p, sev_map in by_pillar.items()
    }
    return AggregatedReport(
        assessment_id=_ASSESSMENT_ID,
        tenant_id=_TENANT_ID,
        total_resources=20,
        resources_with_findings=len({f.resource_id for f in findings}),
        total_findings=len(findings),
        findings_by_pillar=pillars,
        findings_by_severity=by_sev,
        top_critical_findings=[f for f in findings if f.severity == Severity.CRITICAL][:5],
        coverage_percentage=0.25,
        generated_at=datetime.now(UTC),
    )


def _make_report() -> AssessmentReport:
    return AssessmentReport(
        id=_REPORT_ID,
        assessment_id=_ASSESSMENT_ID,
        tenant_id=_TENANT_ID,
        xlsx_blob_path=_XLSX_PATH,
        pdf_blob_path=_PDF_PATH,
        summary=AssessmentSummary(
            assessment_id=_ASSESSMENT_ID,
            tenant_id=_TENANT_ID,
            total_resources=20,
            total_findings=5,
            findings_by_severity={"medium": 5},
            findings_by_pillar={"security": 3, "reliability": 2},
            coverage_percentage=0.25,
        ),
        generated_at=datetime.now(UTC),
    )


def _make_raw_event(total_findings: int = 5) -> bytes:
    event = ReportingRequestedEvent(
        assessment_id=_ASSESSMENT_ID,
        tenant_id=_TENANT_ID,
        batch_id=_BATCH_ID,
        total_findings=total_findings,
    )
    return CloudEventEnvelope.wrap(
        event_type="com.wafagent.reporting.requested",
        source="/agents/reasoning",
        data=event,
    ).to_json_bytes()


def _build_handler(
    *,
    assessment: Assessment | None = None,
    findings: list[Finding] | None = None,
    aggregated: AggregatedReport | None = None,
    report: AssessmentReport | None = None,
    webhook_endpoint: TenantWebhookEndpoint | None = None,
    kv_secret: str = "secret-bytes",
    use_real_generators: bool = False,
) -> tuple[ReportingHandler, dict]:
    if findings is None:
        findings = _make_findings(5)
    if assessment is None:
        assessment = _make_assessment()
    if aggregated is None:
        aggregated = _make_aggregated(findings)
    if report is None:
        report = _make_report()

    assessment_repo = MagicMock()
    assessment_repo.get_by_id = AsyncMock(return_value=assessment)
    assessment_repo.update_status = AsyncMock(return_value=assessment)

    finding_repo = MagicMock()
    finding_repo.list_by_assessment = AsyncMock(return_value=findings)

    report_repo = MagicMock()
    report_repo.create = AsyncMock(return_value=report)

    webhook_repo = MagicMock()
    webhook_repo.get_endpoint_by_tenant = AsyncMock(return_value=webhook_endpoint)

    agg_mock = MagicMock(spec=FindingAggregator)
    agg_mock.aggregate = AsyncMock(return_value=aggregated)

    if use_real_generators:
        excel_gen = ExcelGenerator()
        pdf_gen = PdfGenerator()
    else:
        excel_gen = MagicMock(spec=ExcelGenerator)
        excel_gen.generate = MagicMock(return_value=b"PK_mock_xlsx")
        pdf_gen = MagicMock(spec=PdfGenerator)
        pdf_gen.generate = MagicMock(return_value=b"%PDF-mock")

    uploader = MagicMock(spec=StorageUploader)
    uploader.upload_report = AsyncMock(side_effect=[_XLSX_PATH, _PDF_PATH])

    webhook_svc = MagicMock(spec=WebhookService)
    webhook_svc.deliver = AsyncMock(return_value=None)

    kv_client = MagicMock()
    kv_client.get_secret = AsyncMock(return_value=kv_secret)

    logger = MagicMock()
    logger.bind = MagicMock(return_value=logger)
    for method in ("info", "warning", "error", "debug"):
        setattr(logger, method, MagicMock())

    handler = ReportingHandler(
        assessment_repo=assessment_repo,
        finding_repo=finding_repo,
        report_repo=report_repo,
        webhook_repo=webhook_repo,
        aggregator=agg_mock,
        excel_gen=excel_gen,
        pdf_gen=pdf_gen,
        uploader=uploader,
        webhook_service=webhook_svc,
        kv_client=kv_client,
        logger=logger,
    )
    mocks = dict(
        assessment_repo=assessment_repo,
        finding_repo=finding_repo,
        report_repo=report_repo,
        webhook_repo=webhook_repo,
        aggregator=agg_mock,
        excel_gen=excel_gen,
        pdf_gen=pdf_gen,
        uploader=uploader,
        webhook_svc=webhook_svc,
        kv_client=kv_client,
        logger=logger,
    )
    return handler, mocks


# ── Integration: happy path ───────────────────────────────────────────────────

class TestReportingIntegrationHappyPath:
    @pytest.mark.asyncio
    async def test_end_to_end_without_webhook(self) -> None:
        handler, mocks = _build_handler(webhook_endpoint=None)
        await handler.process(_make_raw_event())

        # Assessment transitions to COMPLETED.
        mocks["assessment_repo"].update_status.assert_called_once_with(
            _TENANT_ID, _ASSESSMENT_ID, AssessmentStatus.COMPLETED
        )
        # Report record created.
        mocks["report_repo"].create.assert_called_once()
        # Both files uploaded.
        assert mocks["uploader"].upload_report.call_count == 2

    @pytest.mark.asyncio
    async def test_end_to_end_with_webhook(self) -> None:
        endpoint = TenantWebhookEndpoint(
            id=uuid.uuid4(),
            tenant_id=_TENANT_ID,
            webhook_url="https://hooks.example.com/waf",
            secret_kv_name="kv-wh-secret",
            is_active=True,
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )
        handler, mocks = _build_handler(webhook_endpoint=endpoint, kv_secret="sec-bytes")
        await handler.process(_make_raw_event())

        mocks["kv_client"].get_secret.assert_called_once_with("kv-wh-secret")
        mocks["webhook_svc"].deliver.assert_called_once()
        call_kwargs = mocks["webhook_svc"].deliver.call_args[1]
        assert call_kwargs["webhook_secret"] == b"sec-bytes"

    @pytest.mark.asyncio
    async def test_cloudevent_fields_parsed(self) -> None:
        """CloudEventEnvelope must be deserialized and routed by assessment_id."""
        handler, mocks = _build_handler(webhook_endpoint=None)
        await handler.process(_make_raw_event(total_findings=42))

        mocks["assessment_repo"].get_by_id.assert_called_once_with(
            _TENANT_ID, _ASSESSMENT_ID
        )

    @pytest.mark.asyncio
    async def test_excel_generator_called_with_real_findings(self) -> None:
        findings = _make_findings(10)
        handler, mocks = _build_handler(findings=findings, webhook_endpoint=None)
        mocks["finding_repo"].list_by_assessment = AsyncMock(return_value=findings)
        await handler.process(_make_raw_event(total_findings=10))

        call_args = mocks["excel_gen"].generate.call_args
        passed_findings = call_args[0][1]
        assert len(passed_findings) == 10


# ── Integration: real generator smoke tests ───────────────────────────────────

class TestRealGeneratorSmoke:
    @pytest.mark.asyncio
    async def test_real_excel_bytes_valid_xlsx(self) -> None:
        import openpyxl

        findings = _make_findings(8)
        handler, mocks = _build_handler(
            findings=findings,
            webhook_endpoint=None,
            use_real_generators=True,
        )
        await handler.process(_make_raw_event(total_findings=8))

        # The uploader received the real Excel bytes.
        xlsx_call = mocks["uploader"].upload_report.call_args_list[0]
        xlsx_data = xlsx_call[1]["data"]
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
        assert "Executive Summary" in wb.sheetnames
        assert "All Findings" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_real_pdf_bytes_valid_pdf(self) -> None:
        findings = _make_findings(5)
        handler, mocks = _build_handler(
            findings=findings,
            webhook_endpoint=None,
            use_real_generators=True,
        )
        await handler.process(_make_raw_event(total_findings=5))

        pdf_call = mocks["uploader"].upload_report.call_args_list[1]
        pdf_data = pdf_call[1]["data"]
        assert pdf_data[:4] == b"%PDF"


# ── Integration: status guards ────────────────────────────────────────────────

class TestReportingIntegrationStatusGuards:
    @pytest.mark.asyncio
    @pytest.mark.parametrize(
        "status",
        [
            AssessmentStatus.COMPLETED,
            AssessmentStatus.FAILED,
            AssessmentStatus.CANCELLED,
            AssessmentStatus.PARTIAL_FAILURE,
        ],
    )
    async def test_terminal_statuses_short_circuit(self, status: AssessmentStatus) -> None:
        handler, mocks = _build_handler(
            assessment=_make_assessment(status),
            webhook_endpoint=None,
        )
        await handler.process(_make_raw_event())

        mocks["aggregator"].aggregate.assert_not_called()
        mocks["report_repo"].create.assert_not_called()
        mocks["assessment_repo"].update_status.assert_not_called()

    @pytest.mark.asyncio
    async def test_none_assessment_short_circuits(self) -> None:
        handler, mocks = _build_handler(webhook_endpoint=None)
        mocks["assessment_repo"].get_by_id = AsyncMock(return_value=None)
        await handler.process(_make_raw_event())

        mocks["aggregator"].aggregate.assert_not_called()

    @pytest.mark.asyncio
    async def test_wrong_status_not_reporting_skips(self) -> None:
        handler, mocks = _build_handler(
            assessment=_make_assessment(AssessmentStatus.EXTRACTING),
            webhook_endpoint=None,
        )
        await handler.process(_make_raw_event())

        mocks["aggregator"].aggregate.assert_not_called()


# ── Integration: error isolation ─────────────────────────────────────────────

class TestReportingIntegrationErrorIsolation:
    @pytest.mark.asyncio
    async def test_kv_error_swallowed_assessment_completed(self) -> None:
        endpoint = TenantWebhookEndpoint(
            id=uuid.uuid4(),
            tenant_id=_TENANT_ID,
            webhook_url="https://hooks.example.com/waf",
            secret_kv_name="bad-secret",
            is_active=True,
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )
        handler, mocks = _build_handler(webhook_endpoint=endpoint)
        mocks["kv_client"].get_secret = AsyncMock(side_effect=RuntimeError("KV gone"))

        await handler.process(_make_raw_event())

        mocks["assessment_repo"].update_status.assert_called_once_with(
            _TENANT_ID, _ASSESSMENT_ID, AssessmentStatus.COMPLETED
        )

    @pytest.mark.asyncio
    async def test_webhook_error_swallowed_assessment_completed(self) -> None:
        endpoint = TenantWebhookEndpoint(
            id=uuid.uuid4(),
            tenant_id=_TENANT_ID,
            webhook_url="https://hooks.example.com/waf",
            secret_kv_name="good-secret",
            is_active=True,
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )
        handler, mocks = _build_handler(webhook_endpoint=endpoint)
        mocks["webhook_svc"].deliver = AsyncMock(
            side_effect=WebhookDeliveryError("https://hooks.example.com/waf", 4)
        )

        await handler.process(_make_raw_event())

        mocks["assessment_repo"].update_status.assert_called_once_with(
            _TENANT_ID, _ASSESSMENT_ID, AssessmentStatus.COMPLETED
        )

    @pytest.mark.asyncio
    async def test_storage_error_propagates(self) -> None:
        """Storage upload failure propagates so Service Bus retries the message."""
        handler, mocks = _build_handler(webhook_endpoint=None)
        mocks["uploader"].upload_report = AsyncMock(side_effect=Exception("Blob down"))

        with pytest.raises(Exception, match="Blob down"):
            await handler.process(_make_raw_event())

        # Assessment must NOT be marked completed when storage fails.
        mocks["assessment_repo"].update_status.assert_not_called()


# ── Integration: report field validation ─────────────────────────────────────

class TestReportFieldValidation:
    @pytest.mark.asyncio
    async def test_report_has_correct_assessment_and_tenant(self) -> None:
        handler, mocks = _build_handler(webhook_endpoint=None)
        await handler.process(_make_raw_event())

        created: AssessmentReport = mocks["report_repo"].create.call_args[0][0]
        assert created.assessment_id == _ASSESSMENT_ID
        assert created.tenant_id == _TENANT_ID

    @pytest.mark.asyncio
    async def test_report_blob_paths_follow_convention(self) -> None:
        handler, mocks = _build_handler(webhook_endpoint=None)
        await handler.process(_make_raw_event())

        created: AssessmentReport = mocks["report_repo"].create.call_args[0][0]
        assert f"reports/{_TENANT_ID}/{_ASSESSMENT_ID}" in created.xlsx_blob_path
        assert f"reports/{_TENANT_ID}/{_ASSESSMENT_ID}" in created.pdf_blob_path
        # Ensure SAS tokens are absent.
        assert "?" not in created.xlsx_blob_path
        assert "?" not in created.pdf_blob_path

    @pytest.mark.asyncio
    async def test_upload_uses_correct_content_types(self) -> None:
        handler, mocks = _build_handler(webhook_endpoint=None)
        await handler.process(_make_raw_event())

        calls = mocks["uploader"].upload_report.call_args_list
        assert len(calls) == 2
        xlsx_call = calls[0]
        pdf_call = calls[1]
        assert xlsx_call[1]["extension"] == "xlsx"
        assert "spreadsheetml" in xlsx_call[1]["content_type"]
        assert pdf_call[1]["extension"] == "pdf"
        assert pdf_call[1]["content_type"] == "application/pdf"

    @pytest.mark.asyncio
    async def test_event_source_and_type_irrelevant_to_handler(self) -> None:
        """Handler processes the event regardless of envelope source/type."""
        event = ReportingRequestedEvent(
            assessment_id=_ASSESSMENT_ID,
            tenant_id=_TENANT_ID,
            batch_id=_BATCH_ID,
            total_findings=1,
        )
        envelope = CloudEventEnvelope.wrap(
            event_type="some.other.type",
            source="/different/source",
            data=event,
        )
        raw = envelope.to_json_bytes()

        handler, mocks = _build_handler(webhook_endpoint=None)
        await handler.process(raw)

        mocks["report_repo"].create.assert_called_once()
