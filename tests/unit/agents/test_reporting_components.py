"""Unit tests for reporting agent components.

Tests aggregator, Excel generator, PDF generator, storage uploader, and
webhook service in isolation — all DB / HTTP / Azure calls are mocked.
"""

from __future__ import annotations

import hashlib
import hmac
import io
import json
import uuid
from datetime import UTC, datetime
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from waf_reporting.aggregator import (
    AggregatedReport,
    FindingAggregator,
    PillarSummary,
    _pillar_compliance_score,
)
from waf_reporting.excel_generator import ExcelGenerator
from waf_reporting.pdf_generator import PdfGenerator
from waf_reporting.storage_uploader import StorageUploader, StorageUploadError
from waf_reporting.webhook_service import WebhookDeliveryError, WebhookService

from waf_shared.domain.models.finding import Finding, FindingStatus, Severity
from waf_shared.domain.models.webhook import WebhookDelivery

# ── Shared fixtures ────────────────────────────────────────────────────────────

_TENANT_ID = uuid.UUID("aaaa0000-0000-0000-0000-000000000001")
_ASSESSMENT_ID = uuid.UUID("bbbb0000-0000-0000-0000-000000000002")
_BATCH_ID = uuid.UUID("cccc0000-0000-0000-0000-000000000003")


def _make_finding(
    severity: str = "medium",
    pillar: str = "security",
    resource_id: str = "/subscriptions/sub-1/resourceGroups/rg-prod/providers/Some/resource",
    waf_codes: list[str] | None = None,
) -> Finding:
    return Finding(
        id=uuid.uuid4(),
        assessment_id=_ASSESSMENT_ID,
        batch_id=_BATCH_ID,
        tenant_id=_TENANT_ID,
        rule_id="WAF-SEC-001",
        resource_id=resource_id,
        resource_type="Microsoft.Network/applicationGateways",
        status=FindingStatus.OPEN,
        severity=Severity(severity),
        pillar=pillar,
        confidence_score=0.9,
        title="Test finding",
        recommendation="Fix it",
        evidence={"result": "FAIL"},
        evaluation_type="deterministic",
        created_at=datetime.now(UTC),
        waf_codes=waf_codes or [],
    )


def _make_aggregated(
    total_findings: int = 3,
    pillars: dict[str, PillarSummary] | None = None,
) -> AggregatedReport:
    if pillars is None:
        pillars = {
            "security": PillarSummary(
                pillar="security",
                findings_by_severity={"critical": 1, "medium": 2},
                total_findings=3,
                compliance_score=0.5,
            )
        }
    return AggregatedReport(
        assessment_id=_ASSESSMENT_ID,
        tenant_id=_TENANT_ID,
        total_resources=10,
        resources_with_findings=3,
        total_findings=total_findings,
        findings_by_pillar=pillars,
        findings_by_severity={"critical": 1, "medium": 2},
        top_critical_findings=[_make_finding("critical")],
        coverage_percentage=0.3,
        generated_at=datetime.now(UTC),
    )


# ── Aggregator tests ──────────────────────────────────────────────────────────


class TestPillarComplianceScore:
    def test_all_critical_gives_zero(self) -> None:
        score = _pillar_compliance_score({"critical": 5})
        assert score == 0.0

    def test_all_informational_gives_one(self) -> None:
        score = _pillar_compliance_score({"informational": 10})
        assert score == 1.0

    def test_empty_gives_one(self) -> None:
        score = _pillar_compliance_score({})
        assert score == 1.0

    def test_mixed_severity_between_zero_and_one(self) -> None:
        score = _pillar_compliance_score({"critical": 2, "low": 2})
        assert 0.0 < score < 1.0

    def test_precision_capped_at_four_decimals(self) -> None:
        score = _pillar_compliance_score({"high": 3, "medium": 1})
        assert len(str(score).split(".")[-1]) <= 4


class TestFindingAggregator:
    def _build_aggregator(
        self,
        by_severity: dict[str, int] | None = None,
        pillar_severity: dict[str, dict[str, int]] | None = None,
        total_resources: int = 10,
        resources_with_findings: int = 3,
        top_critical: list[Finding] | None = None,
    ) -> tuple[FindingAggregator, MagicMock, MagicMock]:
        finding_repo = MagicMock()
        finding_repo.count_by_severity = AsyncMock(
            return_value=by_severity or {"critical": 1, "medium": 2}
        )
        finding_repo.aggregate_pillar_severity = AsyncMock(
            return_value=pillar_severity or {"security": {"critical": 1, "medium": 2}}
        )
        finding_repo.count_distinct_resources = AsyncMock(return_value=resources_with_findings)
        finding_repo.list_by_assessment = AsyncMock(
            return_value=top_critical or [_make_finding("critical")]
        )
        # Phase 5 methods — return empty results so scoring still runs cleanly
        finding_repo.aggregate_resource_type_severity = AsyncMock(return_value={})
        finding_repo.list_top_risks = AsyncMock(return_value=[])
        finding_repo.aggregate_waf_control_coverage = AsyncMock(return_value={})

        assessment_repo = MagicMock()
        assessment_repo.count_resources = AsyncMock(return_value=total_resources)
        assessment_repo.aggregate_resource_inventory = AsyncMock(return_value=[])

        agg = FindingAggregator(
            finding_repo=finding_repo,
            assessment_repo=assessment_repo,
        )
        return agg, finding_repo, assessment_repo

    @pytest.mark.asyncio
    async def test_aggregate_returns_aggregated_report(self) -> None:
        agg, _, _ = self._build_aggregator()
        result = await agg.aggregate(_TENANT_ID, _ASSESSMENT_ID)
        assert isinstance(result, AggregatedReport)
        assert result.assessment_id == _ASSESSMENT_ID
        assert result.tenant_id == _TENANT_ID

    @pytest.mark.asyncio
    async def test_total_findings_sum_of_severity_counts(self) -> None:
        agg, _, _ = self._build_aggregator(by_severity={"critical": 2, "high": 3, "medium": 5})
        result = await agg.aggregate(_TENANT_ID, _ASSESSMENT_ID)
        assert result.total_findings == 10

    @pytest.mark.asyncio
    async def test_coverage_percentage(self) -> None:
        agg, _, _ = self._build_aggregator(total_resources=20, resources_with_findings=5)
        result = await agg.aggregate(_TENANT_ID, _ASSESSMENT_ID)
        assert result.coverage_percentage == pytest.approx(0.25, abs=1e-4)

    @pytest.mark.asyncio
    async def test_zero_resources_coverage_is_zero(self) -> None:
        agg, _, _ = self._build_aggregator(total_resources=0, resources_with_findings=0)
        result = await agg.aggregate(_TENANT_ID, _ASSESSMENT_ID)
        assert result.coverage_percentage == 0.0

    @pytest.mark.asyncio
    async def test_pillar_summaries_built(self) -> None:
        agg, _, _ = self._build_aggregator(
            pillar_severity={
                "security": {"critical": 1},
                "reliability": {"medium": 2},
            }
        )
        result = await agg.aggregate(_TENANT_ID, _ASSESSMENT_ID)
        assert "security" in result.findings_by_pillar
        assert "reliability" in result.findings_by_pillar
        assert result.findings_by_pillar["security"].total_findings == 1
        assert result.findings_by_pillar["reliability"].total_findings == 2

    @pytest.mark.asyncio
    async def test_top_critical_limited_to_five(self) -> None:
        top_critical = [_make_finding("critical") for _ in range(5)]
        agg, finding_repo, _ = self._build_aggregator(top_critical=top_critical)
        result = await agg.aggregate(_TENANT_ID, _ASSESSMENT_ID)
        # list_by_assessment was called with limit=5 for critical findings.
        call_kwargs = finding_repo.list_by_assessment.call_args[1]
        assert call_kwargs["limit"] == 5
        assert call_kwargs["severity"] == Severity.CRITICAL

    @pytest.mark.asyncio
    async def test_generated_at_is_recent(self) -> None:
        agg, _, _ = self._build_aggregator()
        before = datetime.now(UTC)
        result = await agg.aggregate(_TENANT_ID, _ASSESSMENT_ID)
        after = datetime.now(UTC)
        assert before <= result.generated_at <= after


# ── Excel generator tests ─────────────────────────────────────────────────────


class TestExcelGenerator:
    def _gen(self) -> ExcelGenerator:
        return ExcelGenerator()

    def test_returns_bytes(self) -> None:
        findings = [_make_finding("critical"), _make_finding("medium")]
        result = self._gen().generate(_make_aggregated(), findings)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_xlsx_magic_bytes(self) -> None:
        result = self._gen().generate(_make_aggregated(), [_make_finding()])
        # OOXML zip starts with PK header.
        assert result[:2] == b"PK"

    def test_empty_findings_does_not_raise(self) -> None:
        agg = _make_aggregated(total_findings=0, pillars={})
        result = self._gen().generate(agg, [])
        assert isinstance(result, bytes)

    def test_multi_pillar_creates_multiple_sheets(self) -> None:
        import openpyxl

        pillars = {
            "security": PillarSummary(
                pillar="security",
                findings_by_severity={"critical": 1},
                total_findings=1,
                compliance_score=0.0,
            ),
            "reliability": PillarSummary(
                pillar="reliability",
                findings_by_severity={"medium": 2},
                total_findings=2,
                compliance_score=0.5,
            ),
        }
        agg = _make_aggregated(total_findings=3, pillars=pillars)
        findings = [
            _make_finding("critical", "security"),
            _make_finding("medium", "reliability"),
            _make_finding("medium", "reliability"),
        ]
        xlsx_bytes = self._gen().generate(agg, findings)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        sheet_names = wb.sheetnames
        assert "Executive Summary" in sheet_names
        assert "Security" in sheet_names
        assert "Reliability" in sheet_names
        assert "All Findings" in sheet_names

    def test_severity_rows_present_in_all_findings_sheet(self) -> None:
        import openpyxl

        findings = [
            _make_finding("critical"),
            _make_finding("high"),
            _make_finding("low"),
        ]
        xlsx_bytes = self._gen().generate(_make_aggregated(), findings)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["All Findings"]
        values = [[cell.value for cell in row] for row in ws.iter_rows()]
        titles = [row[7] for row in values[1:] if len(row) > 7]  # col H = Title
        assert len(titles) == 3


# ── PDF generator tests ───────────────────────────────────────────────────────


class TestPdfGenerator:
    def _gen(self) -> PdfGenerator:
        return PdfGenerator()

    def test_returns_bytes(self) -> None:
        findings = [_make_finding("critical"), _make_finding("medium")]
        result = self._gen().generate(_make_aggregated(), findings)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_pdf_magic_bytes(self) -> None:
        result = self._gen().generate(_make_aggregated(), [_make_finding()])
        assert result[:4] == b"%PDF"

    def test_empty_findings_does_not_raise(self) -> None:
        agg = _make_aggregated(total_findings=0, pillars={})
        result = self._gen().generate(agg, [])
        assert isinstance(result, bytes)

    def test_many_findings_does_not_raise(self) -> None:
        findings = [_make_finding() for _ in range(200)]
        result = self._gen().generate(_make_aggregated(total_findings=200), findings)
        assert isinstance(result, bytes)

    def test_critical_findings_render_risk_statement(self) -> None:
        findings = [_make_finding("critical") for _ in range(3)]
        agg = _make_aggregated(total_findings=3)
        result = self._gen().generate(agg, findings)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_waf_codes_render_control_pages(self) -> None:
        findings = [
            _make_finding("critical", "security", waf_codes=["SE-05", "SE-08"]),
            _make_finding("high", "security", waf_codes=["SE-05"]),
            _make_finding("medium", "reliability", waf_codes=["RE-01"]),
        ]
        agg = _make_aggregated(total_findings=3)
        result = self._gen().generate(agg, findings)
        assert isinstance(result, bytes)
        assert result[:4] == b"%PDF"

    def test_resource_group_breakdown_parses_arm_ids(self) -> None:
        findings = [
            _make_finding(
                resource_id="/subscriptions/s1/resourceGroups/rg-frontend/providers/Some/r"
            ),
            _make_finding(
                resource_id="/subscriptions/s1/resourceGroups/rg-backend/providers/Some/r"
            ),
            _make_finding(
                "critical",
                resource_id="/subscriptions/s1/resourceGroups/rg-frontend/providers/Some/r2",
            ),
        ]
        agg = _make_aggregated(total_findings=3)
        result = self._gen().generate(agg, findings)
        assert isinstance(result, bytes)

    def test_multi_pillar_all_sections_render(self) -> None:
        pillars = {
            "security": PillarSummary(
                pillar="security",
                findings_by_severity={"critical": 1},
                total_findings=1,
                compliance_score=0.0,
            ),
            "reliability": PillarSummary(
                pillar="reliability",
                findings_by_severity={"medium": 2},
                total_findings=2,
                compliance_score=0.5,
            ),
        }
        findings = [
            _make_finding("critical", "security", waf_codes=["SE-05"]),
            _make_finding("medium", "reliability", waf_codes=["RE-01"]),
            _make_finding("medium", "reliability"),
        ]
        agg = _make_aggregated(total_findings=3, pillars=pillars)
        result = self._gen().generate(agg, findings)
        assert isinstance(result, bytes)
        assert result[:4] == b"%PDF"

    def test_generate_with_human_reviews(self) -> None:
        from waf_shared.domain.models.human_review import (
            ComplianceStatus,
            HumanReviewAssessment,
            ReviewStatus,
        )

        review = HumanReviewAssessment(
            id=uuid.uuid4(),
            assessment_id=_ASSESSMENT_ID,
            tenant_id=_TENANT_ID,
            control_code="SE-10",
            pillar="security",
            reviewer_oid="reviewer-001",
            status=ReviewStatus.COMPLETED,
            compliance_status=ComplianceStatus.COMPLIANT,
            score=85,
            comments="Pen test completed quarterly",
            reviewed_at=datetime.now(UTC),
            created_at=datetime.now(UTC),
            updated_at=datetime.now(UTC),
        )
        findings = [_make_finding("medium")]
        result = self._gen().generate(_make_aggregated(), findings, human_reviews=[review])
        assert isinstance(result, bytes)
        assert result[:4] == b"%PDF"


# ── group_findings_for_reporting tests ────────────────────────────────────────

from waf_reporting.pdf_generator import (  # noqa: E402
    build_evidence_snapshot,
    build_executive_remediation_roadmap,
    calculate_maturity_rating,
    calculate_pillar_scores,
    calculate_remediation_priority,
    estimate_effort,
    group_findings_for_reporting,
    sanitize_evidence,
)


def _make_finding_for_group(
    rule_id: str = "WAF-SEC-001",
    severity: str = "high",
    resource_id: str = "/subscriptions/sub/resourceGroups/rg/providers/Storage/accounts/st1",
    recommendation: str = "Enable secure transfer.",
    pillar: str = "security",
    waf_codes: list[str] | None = None,
    evidence: dict | None = None,
) -> Finding:
    return Finding(
        id=uuid.uuid4(),
        assessment_id=_ASSESSMENT_ID,
        batch_id=_BATCH_ID,
        tenant_id=_TENANT_ID,
        rule_id=rule_id,
        resource_id=resource_id,
        resource_type="Microsoft.Storage/storageAccounts",
        status=FindingStatus.OPEN,
        severity=Severity(severity),
        pillar=pillar,
        confidence_score=0.95,
        title="Storage account should require secure transfer",
        recommendation=recommendation,
        evidence=evidence or {"result": "FAIL"},
        evaluation_type="deterministic",
        created_at=datetime.now(UTC),
        waf_codes=waf_codes or ["SE-03"],
    )


class TestGroupFindingsForReporting:
    """Unit tests for the public group_findings_for_reporting() helper."""

    def test_empty_input_returns_empty_list(self) -> None:
        assert group_findings_for_reporting([]) == []

    def test_single_finding_produces_one_group(self) -> None:
        f = _make_finding_for_group()
        groups = group_findings_for_reporting([f])
        assert len(groups) == 1
        assert groups[0].count == 1

    def test_three_same_rule_same_severity_collapse_to_one_group(self) -> None:
        findings = [
            _make_finding_for_group(resource_id=f"/rg/providers/Storage/accounts/st{i}")
            for i in range(1, 4)
        ]
        groups = group_findings_for_reporting(findings)
        assert len(groups) == 1
        assert groups[0].count == 3

    def test_different_severity_stays_in_separate_groups(self) -> None:
        findings = [
            _make_finding_for_group(severity="high"),
            _make_finding_for_group(severity="medium"),
        ]
        groups = group_findings_for_reporting(findings)
        assert len(groups) == 2
        severities = {g.severity for g in groups}
        assert severities == {"high", "medium"}

    def test_different_rule_id_stays_in_separate_groups(self) -> None:
        findings = [
            _make_finding_for_group(rule_id="WAF-SEC-001"),
            _make_finding_for_group(rule_id="WAF-SEC-002"),
        ]
        groups = group_findings_for_reporting(findings)
        assert len(groups) == 2

    def test_resource_names_deduplicated_within_group(self) -> None:
        resource = "/rg/providers/Storage/accounts/stshared"
        findings = [_make_finding_for_group(resource_id=resource) for _ in range(3)]
        groups = group_findings_for_reporting(findings)
        assert len(groups) == 1
        assert groups[0].count == 1
        assert groups[0].resource_names == ["stshared"]

    def test_resource_names_are_short_form_after_last_slash(self) -> None:
        f = _make_finding_for_group(
            resource_id="/subscriptions/sub/resourceGroups/rg/providers/S/accounts/myaccount"
        )
        groups = group_findings_for_reporting([f])
        assert groups[0].resource_names == ["myaccount"]

    def test_resource_id_without_slash_used_as_is(self) -> None:
        f = _make_finding_for_group(resource_id="my-flat-resource")
        groups = group_findings_for_reporting([f])
        assert groups[0].resource_names == ["my-flat-resource"]

    def test_waf_codes_merged_across_findings(self) -> None:
        findings = [
            _make_finding_for_group(resource_id="/rg/s/st1", waf_codes=["SE-03"]),
            _make_finding_for_group(resource_id="/rg/s/st2", waf_codes=["SE-03", "SE-07"]),
        ]
        groups = group_findings_for_reporting(findings)
        assert len(groups) == 1
        codes = groups[0].waf_codes
        assert "SE-03" in codes
        assert "SE-07" in codes

    def test_most_voted_recommendation_is_selected(self) -> None:
        findings = [
            _make_finding_for_group(resource_id=f"/rg/s/st{i}", recommendation="Enable HTTPS.")
            for i in range(3)
        ] + [
            _make_finding_for_group(resource_id="/rg/s/st99", recommendation="Use secure transfer.")
        ]
        groups = group_findings_for_reporting(findings)
        assert len(groups) == 1
        assert groups[0].recommendation == "Enable HTTPS."

    def test_evidence_summary_populated(self) -> None:
        findings = [
            _make_finding_for_group(resource_id=f"/rg/s/st{i}", evidence={"result": "FAIL"})
            for i in range(2)
        ]
        groups = group_findings_for_reporting(findings)
        assert len(groups) == 1
        assert groups[0].evidence_summary != ""
        assert "2" in groups[0].evidence_summary

    def test_evidence_summary_with_no_evidence_dict(self) -> None:
        f = _make_finding_for_group(evidence={})
        groups = group_findings_for_reporting([f])
        assert groups[0].evidence_summary != ""

    def test_sorted_worst_severity_first(self) -> None:
        findings = [
            _make_finding_for_group(rule_id="R1", severity="low"),
            _make_finding_for_group(rule_id="R2", severity="critical"),
            _make_finding_for_group(rule_id="R3", severity="medium"),
        ]
        groups = group_findings_for_reporting(findings)
        assert groups[0].severity == "critical"
        assert groups[-1].severity == "low"

    def test_sorted_by_count_descending_within_same_severity(self) -> None:
        findings = [
            _make_finding_for_group(rule_id="R1", resource_id="/rg/s/a1"),
            _make_finding_for_group(rule_id="R2", resource_id="/rg/s/b1"),
            _make_finding_for_group(rule_id="R2", resource_id="/rg/s/b2"),
        ]
        groups = group_findings_for_reporting(findings)
        # R2 has 2 resources → should sort first within same severity
        assert groups[0].rule_id == "R2"
        assert groups[1].rule_id == "R1"

    def test_same_rule_different_recommendation_merges_into_one_group(self) -> None:
        """Core deduplication test: same rule + severity, different recommendation text."""
        findings = [
            _make_finding_for_group(
                resource_id=f"/rg/s/st{i}",
                recommendation=f"Fix recommendation variant {i}",
            )
            for i in range(4)
        ]
        groups = group_findings_for_reporting(findings)
        # All 4 must collapse into a single group — this is the key fix.
        assert len(groups) == 1
        assert groups[0].count == 4

    def test_pdf_generates_successfully_with_grouped_findings(self) -> None:
        """Integration smoke test: PDF generator uses group_findings_for_reporting."""
        findings = [_make_finding_for_group(resource_id=f"/rg/s/st{i}") for i in range(4)]
        agg = _make_aggregated(
            total_findings=4,
            pillars={
                "security": PillarSummary(
                    pillar="security",
                    findings_by_severity={"high": 4},
                    total_findings=4,
                    compliance_score=0.25,
                )
            },
        )
        result = PdfGenerator().generate(agg, findings)
        assert isinstance(result, bytes)
        assert result[:4] == b"%PDF"


# ── Storage uploader tests ─────────────────────────────────────────────────────


class TestStorageUploader:
    def _build_uploader(self, upload_error: Exception | None = None) -> StorageUploader:
        blob_client = AsyncMock()
        blob_client.upload_blob = AsyncMock(side_effect=upload_error if upload_error else None)
        container_client = MagicMock()
        container_client.get_blob_client = MagicMock(return_value=blob_client)
        blob_service = MagicMock()
        blob_service.get_container_client = MagicMock(return_value=container_client)
        logger = MagicMock()
        logger.bind = MagicMock(return_value=logger)
        logger.info = MagicMock()
        logger.error = MagicMock()
        return StorageUploader(
            blob_service=blob_service, container_name="waf-reports", logger=logger
        )

    @pytest.mark.asyncio
    async def test_returns_blob_path_not_sas_url(self) -> None:
        uploader = self._build_uploader()
        path = await uploader.upload_report(
            tenant_id=_TENANT_ID,
            assessment_id=_ASSESSMENT_ID,
            data=b"xlsx_data",
            extension="xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        assert path == f"reports/{_TENANT_ID}/{_ASSESSMENT_ID}/report.xlsx"
        assert "?" not in path
        assert "sig=" not in path

    @pytest.mark.asyncio
    async def test_pdf_blob_path(self) -> None:
        uploader = self._build_uploader()
        path = await uploader.upload_report(
            tenant_id=_TENANT_ID,
            assessment_id=_ASSESSMENT_ID,
            data=b"pdf_data",
            extension="pdf",
            content_type="application/pdf",
        )
        assert path.endswith(".pdf")
        assert str(_TENANT_ID) in path
        assert str(_ASSESSMENT_ID) in path

    @pytest.mark.asyncio
    async def test_upload_error_raises_storage_upload_error(self) -> None:
        uploader = self._build_uploader(upload_error=Exception("Azure SDK failure"))
        with pytest.raises(StorageUploadError):
            await uploader.upload_report(
                tenant_id=_TENANT_ID,
                assessment_id=_ASSESSMENT_ID,
                data=b"data",
                extension="xlsx",
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


# ── Webhook service tests ──────────────────────────────────────────────────────


class TestWebhookService:
    def _build_service(
        self,
        http_responses: list[int] | None = None,
        http_errors: list[Exception | None] | None = None,
    ) -> tuple[WebhookService, MagicMock]:
        webhook_repo = MagicMock()
        webhook_repo.record_delivery = AsyncMock(return_value=MagicMock())

        logger = MagicMock()
        logger.bind = MagicMock(return_value=logger)
        logger.info = MagicMock()
        logger.warning = MagicMock()
        logger.error = MagicMock()
        logger.debug = MagicMock()

        svc = WebhookService(webhook_repo=webhook_repo, logger=logger)
        return svc, webhook_repo

    def _patch_aiohttp(self, status_codes: list[int], errors: list[Exception | None] | None = None):
        responses = []
        err_list = errors or [None] * len(status_codes)
        for code, err in zip(status_codes, err_list, strict=False):
            if err:
                responses.append(err)
            else:
                mock_resp = AsyncMock()
                mock_resp.status = code
                mock_resp.__aenter__ = AsyncMock(return_value=mock_resp)
                mock_resp.__aexit__ = AsyncMock(return_value=False)
                responses.append(mock_resp)

        mock_session = MagicMock()
        mock_session.__aenter__ = AsyncMock(return_value=mock_session)
        mock_session.__aexit__ = AsyncMock(return_value=False)

        call_count = [0]

        async def _post(url, **kwargs):
            idx = call_count[0]
            call_count[0] += 1
            r = responses[idx] if idx < len(responses) else responses[-1]
            if isinstance(r, Exception):
                raise r
            return r

        mock_session.post = _post
        return mock_session

    @pytest.mark.asyncio
    async def test_successful_delivery_first_attempt(self) -> None:
        svc, repo = self._build_service()
        mock_resp = AsyncMock()
        mock_resp.status = 200
        mock_resp.__aenter__ = AsyncMock(return_value=mock_resp)
        mock_resp.__aexit__ = AsyncMock(return_value=False)

        with patch("aiohttp.ClientSession") as mock_cls:
            session = MagicMock()
            session.__aenter__ = AsyncMock(return_value=session)
            session.__aexit__ = AsyncMock(return_value=False)
            session.post = MagicMock(return_value=mock_resp)
            mock_cls.return_value = session

            await svc.deliver(
                tenant_id=_TENANT_ID,
                assessment_id=_ASSESSMENT_ID,
                webhook_url="https://example.com/hook",
                webhook_secret=b"secret",
                payload={"status": "completed"},
            )

        repo.record_delivery.assert_called_once()
        delivery_arg: WebhookDelivery = repo.record_delivery.call_args[0][0]
        assert delivery_arg.success is True
        assert delivery_arg.status_code == 200

    @pytest.mark.asyncio
    async def test_hmac_signature_header_computed(self) -> None:
        """HMAC-SHA256 signature in delivery body must match the expected value."""
        captured_headers: dict[str, str] = {}
        captured_body: bytes = b""

        mock_resp = AsyncMock()
        mock_resp.status = 200
        mock_resp.__aenter__ = AsyncMock(return_value=mock_resp)
        mock_resp.__aexit__ = AsyncMock(return_value=False)

        async def fake_post(url, *, data, headers, timeout):
            nonlocal captured_headers, captured_body
            captured_headers = dict(headers)
            captured_body = data
            return mock_resp

        svc, _ = self._build_service()
        with patch("aiohttp.ClientSession") as mock_cls:
            session = MagicMock()
            session.__aenter__ = AsyncMock(return_value=session)
            session.__aexit__ = AsyncMock(return_value=False)
            session.post = fake_post
            mock_cls.return_value = session

            secret = b"my-webhook-secret"
            payload = {"assessment_id": str(_ASSESSMENT_ID), "status": "completed"}
            await svc.deliver(
                tenant_id=_TENANT_ID,
                assessment_id=_ASSESSMENT_ID,
                webhook_url="https://example.com/hook",
                webhook_secret=secret,
                payload=payload,
            )

        # Verify the HMAC header matches independent calculation.
        expected_sig = hmac.new(secret, captured_body, hashlib.sha256).hexdigest()
        assert captured_headers["X-WAF-Signature"] == f"sha256={expected_sig}"

    @pytest.mark.asyncio
    async def test_all_retries_exhausted_raises_delivery_error(self) -> None:
        svc, repo = self._build_service()

        mock_resp = AsyncMock()
        mock_resp.status = 503
        mock_resp.__aenter__ = AsyncMock(return_value=mock_resp)
        mock_resp.__aexit__ = AsyncMock(return_value=False)

        with (
            patch("aiohttp.ClientSession") as mock_cls,
            patch("asyncio.sleep", new_callable=AsyncMock),
        ):
            session = MagicMock()
            session.__aenter__ = AsyncMock(return_value=session)
            session.__aexit__ = AsyncMock(return_value=False)
            session.post = MagicMock(return_value=mock_resp)
            mock_cls.return_value = session

            with pytest.raises(WebhookDeliveryError):
                await svc.deliver(
                    tenant_id=_TENANT_ID,
                    assessment_id=_ASSESSMENT_ID,
                    webhook_url="https://example.com/hook",
                    webhook_secret=b"secret",
                    payload={"status": "completed"},
                )

        # All 4 attempts logged to DB.
        assert repo.record_delivery.call_count == 4

    @pytest.mark.asyncio
    async def test_delivery_log_failure_does_not_abort_delivery(self) -> None:
        """A DB failure while logging delivery must not prevent retrying."""
        svc, repo = self._build_service()
        repo.record_delivery = AsyncMock(side_effect=Exception("DB down"))

        mock_resp = AsyncMock()
        mock_resp.status = 200
        mock_resp.__aenter__ = AsyncMock(return_value=mock_resp)
        mock_resp.__aexit__ = AsyncMock(return_value=False)

        with patch("aiohttp.ClientSession") as mock_cls:
            session = MagicMock()
            session.__aenter__ = AsyncMock(return_value=session)
            session.__aexit__ = AsyncMock(return_value=False)
            session.post = MagicMock(return_value=mock_resp)
            mock_cls.return_value = session

            # Should complete without raising despite DB failure.
            await svc.deliver(
                tenant_id=_TENANT_ID,
                assessment_id=_ASSESSMENT_ID,
                webhook_url="https://example.com/hook",
                webhook_secret=b"secret",
                payload={"status": "completed"},
            )


# ── calculate_pillar_scores tests ─────────────────────────────────────────────


class TestCalculatePillarScores:
    def test_empty_findings_returns_five_pillars_all_at_100(self) -> None:
        scores = calculate_pillar_scores([])
        assert len(scores) == 5
        assert all(s[1] == 100 for s in scores)
        assert all(s[2] == "Excellent" for s in scores)

    def test_always_returns_five_pillars_regardless_of_input(self) -> None:
        scores = calculate_pillar_scores([_make_finding(severity="high", pillar="security")])
        assert len(scores) == 5

    def test_critical_deducts_15(self) -> None:
        scores = calculate_pillar_scores([_make_finding(severity="critical", pillar="security")])
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 85

    def test_high_deducts_10(self) -> None:
        scores = calculate_pillar_scores([_make_finding(severity="high", pillar="reliability")])
        rel = next(s for s in scores if s[0] == "Reliability")
        assert rel[1] == 90

    def test_medium_deducts_5(self) -> None:
        scores = calculate_pillar_scores(
            [_make_finding(severity="medium", pillar="cost_optimization")]
        )
        co = next(s for s in scores if s[0] == "Cost Optimization")
        assert co[1] == 95

    def test_low_deducts_2(self) -> None:
        scores = calculate_pillar_scores([_make_finding(severity="low", pillar="security")])
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 98

    def test_informational_no_deduction(self) -> None:
        scores = calculate_pillar_scores(
            [_make_finding(severity="informational", pillar="security")]
        )
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 100

    def test_mixed_severities_summed(self) -> None:
        findings = [
            _make_finding(severity="critical", pillar="security"),
            _make_finding(severity="high", pillar="security"),
            _make_finding(severity="medium", pillar="security"),
            _make_finding(severity="low", pillar="security"),
        ]
        scores = calculate_pillar_scores(findings)
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 68  # 100 - (15+10+5+2)

    def test_multiple_pillars_scored_independently(self) -> None:
        findings = [
            _make_finding(severity="critical", pillar="security"),
            _make_finding(severity="high", pillar="reliability"),
        ]
        scores = calculate_pillar_scores(findings)
        sec = next(s for s in scores if s[0] == "Security")
        rel = next(s for s in scores if s[0] == "Reliability")
        assert sec[1] == 85
        assert rel[1] == 90

    def test_score_floor_at_zero(self) -> None:
        findings = [_make_finding(severity="critical", pillar="security") for _ in range(7)]
        scores = calculate_pillar_scores(findings)
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 0  # 100 - 7×15 = -5, floored to 0

    def test_score_ceiling_at_100(self) -> None:
        scores = calculate_pillar_scores(
            [_make_finding(severity="informational", pillar="security")]
        )
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 100

    def test_status_excellent_at_and_above_90(self) -> None:
        # 1 high = -10 → score 90 → Excellent
        scores = calculate_pillar_scores([_make_finding(severity="high", pillar="security")])
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 90
        assert sec[2] == "Excellent"

    def test_status_good_at_75(self) -> None:
        # 2 high + 1 medium = -(20+5) = -25 → score 75 → Good
        findings = [
            _make_finding(severity="high", pillar="security"),
            _make_finding(severity="high", pillar="security"),
            _make_finding(severity="medium", pillar="security"),
        ]
        scores = calculate_pillar_scores(findings)
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 75
        assert sec[2] == "Good"

    def test_status_needs_improvement_at_60(self) -> None:
        # 4 high = -40 → score 60 → Needs Improvement
        findings = [_make_finding(severity="high", pillar="security") for _ in range(4)]
        scores = calculate_pillar_scores(findings)
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 60
        assert sec[2] == "Needs Improvement"

    def test_status_high_risk_below_60(self) -> None:
        # 5 high = -50 → score 50 → High Risk
        findings = [_make_finding(severity="high", pillar="security") for _ in range(5)]
        scores = calculate_pillar_scores(findings)
        sec = next(s for s in scores if s[0] == "Security")
        assert sec[1] == 50
        assert sec[2] == "High Risk"

    def test_pillar_with_no_findings_stays_at_100(self) -> None:
        findings = [_make_finding(severity="critical", pillar="security")]
        scores = calculate_pillar_scores(findings)
        rel = next(s for s in scores if s[0] == "Reliability")
        assert rel[1] == 100
        assert rel[2] == "Excellent"

    def test_finding_counts_correct(self) -> None:
        findings = [
            _make_finding(severity="critical", pillar="security"),
            _make_finding(severity="high", pillar="security"),
            _make_finding(severity="medium", pillar="security"),
        ]
        scores = calculate_pillar_scores(findings)
        sec = next(s for s in scores if s[0] == "Security")
        # tuple: (name, score, status, total, crit, high, med, low)
        assert sec[3] == 3  # total
        assert sec[4] == 1  # critical
        assert sec[5] == 1  # high
        assert sec[6] == 1  # medium
        assert sec[7] == 0  # low


# ── calculate_maturity_rating tests ───────────────────────────────────────────


class TestCalculateMaturityRating:
    def test_enterprise_ready_at_90(self) -> None:
        assert calculate_maturity_rating(90.0) == "Enterprise Ready"

    def test_enterprise_ready_at_100(self) -> None:
        assert calculate_maturity_rating(100.0) == "Enterprise Ready"

    def test_strong_at_80(self) -> None:
        assert calculate_maturity_rating(80.0) == "Strong"

    def test_strong_at_89(self) -> None:
        assert calculate_maturity_rating(89.9) == "Strong"

    def test_moderate_at_70(self) -> None:
        assert calculate_maturity_rating(70.0) == "Moderate"

    def test_moderate_at_79(self) -> None:
        assert calculate_maturity_rating(79.9) == "Moderate"

    def test_needs_improvement_at_60(self) -> None:
        assert calculate_maturity_rating(60.0) == "Needs Improvement"

    def test_needs_improvement_at_69(self) -> None:
        assert calculate_maturity_rating(69.9) == "Needs Improvement"

    def test_high_risk_below_60(self) -> None:
        assert calculate_maturity_rating(59.9) == "High Risk"

    def test_high_risk_at_zero(self) -> None:
        assert calculate_maturity_rating(0.0) == "High Risk"


# ── sanitize_evidence tests ───────────────────────────────────────────────────


class TestSanitizeEvidence:
    def test_empty_evidence_returns_empty(self) -> None:
        assert sanitize_evidence({}) == {}

    def test_non_sensitive_keys_pass_through(self) -> None:
        ev = {"supportsHttpsTrafficOnly": False, "minimumTlsVersion": "TLS1_0"}
        assert sanitize_evidence(ev) == ev

    def test_strips_password_key(self) -> None:
        result = sanitize_evidence({"password": "abc", "status": "fail"})
        assert "password" not in result
        assert result["status"] == "fail"

    def test_strips_secret_key(self) -> None:
        result = sanitize_evidence({"apiSecret": "abc", "count": 1})
        assert "apiSecret" not in result
        assert result["count"] == 1

    def test_strips_token_key(self) -> None:
        result = sanitize_evidence({"accessToken": "xyz", "enabled": True})
        assert "accessToken" not in result
        assert result["enabled"] is True

    def test_strips_key_key(self) -> None:
        result = sanitize_evidence({"storageKey": "abc123", "size": 5})
        assert "storageKey" not in result
        assert result["size"] == 5

    def test_strips_certificate_key(self) -> None:
        result = sanitize_evidence({"sslCertificate": "PEM...", "valid": True})
        assert "sslCertificate" not in result
        assert result["valid"] is True

    def test_strips_sas_key(self) -> None:
        result = sanitize_evidence({"sasToken": "https://...", "count": 3})
        assert "sasToken" not in result
        assert result["count"] == 3

    def test_strips_connectionstring_key(self) -> None:
        result = sanitize_evidence({"connectionString": "Server=...", "timeout": 30})
        assert "connectionString" not in result
        assert result["timeout"] == 30

    def test_case_insensitive_stripping(self) -> None:
        result = sanitize_evidence({"PASSWORD": "abc", "Token": "xyz", "data": "ok"})
        assert "PASSWORD" not in result
        assert "Token" not in result
        assert result["data"] == "ok"

    def test_all_secrets_returns_empty(self) -> None:
        result = sanitize_evidence({"password": "abc", "secret": "xyz"})
        assert result == {}


# ── build_evidence_snapshot tests ─────────────────────────────────────────────


class TestBuildEvidenceSnapshot:
    def _finding_with_evidence(self, evidence: dict) -> Finding:
        return Finding(
            id=uuid.uuid4(),
            assessment_id=_ASSESSMENT_ID,
            batch_id=_BATCH_ID,
            tenant_id=_TENANT_ID,
            rule_id="WAF-SEC-001",
            resource_id="/sub/rg/resource",
            resource_type="Microsoft.Storage/storageAccounts",
            status=FindingStatus.OPEN,
            severity=Severity.HIGH,
            pillar="security",
            confidence_score=0.9,
            title="Test",
            recommendation="Fix it",
            evidence=evidence,
            evaluation_type="deterministic",
            created_at=datetime.now(UTC),
        )

    def test_empty_evidence_returns_empty_dict(self) -> None:
        assert build_evidence_snapshot(self._finding_with_evidence({})) == {}

    def test_normal_evidence_returned(self) -> None:
        snap = build_evidence_snapshot(
            self._finding_with_evidence({"supportsHttpsTrafficOnly": False})
        )
        assert snap == {"supportsHttpsTrafficOnly": False}

    def test_secrets_stripped(self) -> None:
        snap = build_evidence_snapshot(
            self._finding_with_evidence({"password": "abc", "status": "fail"})
        )
        assert "password" not in snap
        assert snap.get("status") == "fail"

    def test_all_secrets_stripped_returns_empty(self) -> None:
        snap = build_evidence_snapshot(
            self._finding_with_evidence({"password": "abc", "secret": "xyz"})
        )
        assert snap == {}

    def test_large_evidence_trimmed_to_10_fields(self) -> None:
        ev = {f"field_{i}": i for i in range(15)}
        snap = build_evidence_snapshot(self._finding_with_evidence(ev))
        assert len(snap) <= 10

    def test_large_serialized_evidence_trimmed_under_500_chars(self) -> None:
        ev = {"key1": "x" * 200, "key2": "y" * 200, "key3": "z" * 200}
        snap = build_evidence_snapshot(self._finding_with_evidence(ev))
        serialized = json.dumps(snap, default=str)
        assert len(serialized) <= 510  # 500 + small truncation marker overhead

    def test_storage_https_finding_evidence(self) -> None:
        snap = build_evidence_snapshot(
            self._finding_with_evidence({"supportsHttpsTrafficOnly": False})
        )
        assert snap["supportsHttpsTrafficOnly"] is False

    def test_tls_version_finding_evidence(self) -> None:
        snap = build_evidence_snapshot(self._finding_with_evidence({"minimumTlsVersion": "TLS1_0"}))
        assert snap["minimumTlsVersion"] == "TLS1_0"

    def test_private_endpoint_finding_evidence(self) -> None:
        snap = build_evidence_snapshot(
            self._finding_with_evidence({"privateEndpointConnections": []})
        )
        assert snap["privateEndpointConnections"] == []

    def test_rbac_finding_evidence(self) -> None:
        snap = build_evidence_snapshot(self._finding_with_evidence({"roleAssignments": 0}))
        assert snap["roleAssignments"] == 0

    def test_multiple_clean_fields_all_returned(self) -> None:
        ev = {
            "supportsHttpsTrafficOnly": False,
            "minimumTlsVersion": "TLS1_0",
            "allowBlobPublicAccess": True,
        }
        snap = build_evidence_snapshot(self._finding_with_evidence(ev))
        assert len(snap) == 3
        assert snap["minimumTlsVersion"] == "TLS1_0"


# ── calculate_remediation_priority tests ──────────────────────────────────────


class TestCalculateRemediationPriority:
    def _make(self, severity: str, pillar: str) -> Finding:
        return Finding(
            id=uuid.uuid4(),
            assessment_id=_ASSESSMENT_ID,
            batch_id=_BATCH_ID,
            tenant_id=_TENANT_ID,
            rule_id="WAF-TEST-001",
            resource_id="/sub/rg/res",
            resource_type="Microsoft.Storage/storageAccounts",
            status=FindingStatus.OPEN,
            severity=Severity(severity),
            pillar=pillar,
            confidence_score=0.9,
            title="Test",
            recommendation="Fix it",
            evidence={},
            evaluation_type="deterministic",
            created_at=datetime.now(UTC),
        )

    def test_critical_security_highest_priority(self) -> None:
        score = calculate_remediation_priority(self._make("critical", "security"))
        assert score == 120  # 100 + 20

    def test_high_security(self) -> None:
        score = calculate_remediation_priority(self._make("high", "security"))
        assert score == 95  # 75 + 20

    def test_medium_reliability(self) -> None:
        score = calculate_remediation_priority(self._make("medium", "reliability"))
        assert score == 65  # 50 + 15

    def test_low_cost_optimization(self) -> None:
        score = calculate_remediation_priority(self._make("low", "cost_optimization"))
        assert score == 35  # 25 + 10

    def test_informational_operational_excellence(self) -> None:
        score = calculate_remediation_priority(
            self._make("informational", "operational_excellence")
        )
        assert score == 10  # 0 + 10

    def test_critical_higher_than_high_same_pillar(self) -> None:
        crit = calculate_remediation_priority(self._make("critical", "reliability"))
        high = calculate_remediation_priority(self._make("high", "reliability"))
        assert crit > high

    def test_same_severity_security_beats_other_pillars(self) -> None:
        sec = calculate_remediation_priority(self._make("high", "security"))
        rel = calculate_remediation_priority(self._make("high", "reliability"))
        assert sec > rel


# ── estimate_effort tests ─────────────────────────────────────────────────────


class TestEstimateEffort:
    def test_zero_resources_returns_low(self) -> None:
        assert estimate_effort(0) == "Low"

    def test_one_resource_returns_low(self) -> None:
        assert estimate_effort(1) == "Low"

    def test_two_resources_returns_medium(self) -> None:
        assert estimate_effort(2) == "Medium"

    def test_five_resources_returns_medium(self) -> None:
        assert estimate_effort(5) == "Medium"

    def test_six_resources_returns_high(self) -> None:
        assert estimate_effort(6) == "High"

    def test_large_count_returns_high(self) -> None:
        assert estimate_effort(100) == "High"


# ── build_executive_remediation_roadmap tests ─────────────────────────────────


class TestBuildExecutiveRemediationRoadmap:
    def _make_finding(
        self,
        severity: str,
        pillar: str,
        rule_id: str = "WAF-SEC-001",
        resource_id: str = "/sub/rg/res",
    ) -> Finding:
        return Finding(
            id=uuid.uuid4(),
            assessment_id=_ASSESSMENT_ID,
            batch_id=_BATCH_ID,
            tenant_id=_TENANT_ID,
            rule_id=rule_id,
            resource_id=resource_id,
            resource_type="Microsoft.Storage/storageAccounts",
            status=FindingStatus.OPEN,
            severity=Severity(severity),
            pillar=pillar,
            confidence_score=0.9,
            title=f"{severity.capitalize()} {pillar} finding",
            recommendation=f"Fix {severity} issue in {pillar}",
            evidence={},
            evaluation_type="deterministic",
            created_at=datetime.now(UTC),
        )

    def test_empty_findings_returns_empty_list(self) -> None:
        assert build_executive_remediation_roadmap([]) == []

    def test_only_critical_produces_phase1_only(self) -> None:
        findings = [self._make_finding("critical", "security")]
        phases = build_executive_remediation_roadmap(findings)
        assert len(phases) == 1
        assert phases[0]["name"] == "Phase 1 — Immediate"

    def test_high_security_goes_to_phase1(self) -> None:
        findings = [self._make_finding("high", "security")]
        phases = build_executive_remediation_roadmap(findings)
        assert len(phases) == 1
        assert phases[0]["name"] == "Phase 1 — Immediate"

    def test_high_non_security_goes_to_phase2(self) -> None:
        findings = [self._make_finding("high", "reliability")]
        phases = build_executive_remediation_roadmap(findings)
        assert len(phases) == 1
        assert phases[0]["name"] == "Phase 2 — Near Term"

    def test_medium_goes_to_phase2(self) -> None:
        findings = [self._make_finding("medium", "operational_excellence")]
        phases = build_executive_remediation_roadmap(findings)
        assert len(phases) == 1
        assert phases[0]["name"] == "Phase 2 — Near Term"

    def test_low_goes_to_phase3(self) -> None:
        findings = [self._make_finding("low", "cost_optimization")]
        phases = build_executive_remediation_roadmap(findings)
        assert len(phases) == 1
        assert phases[0]["name"] == "Phase 3 — Strategic"

    def test_informational_goes_to_phase3(self) -> None:
        findings = [self._make_finding("informational", "performance_efficiency")]
        phases = build_executive_remediation_roadmap(findings)
        assert len(phases) == 1
        assert phases[0]["name"] == "Phase 3 — Strategic"

    def test_mixed_severities_produces_three_phases(self) -> None:
        findings = [
            self._make_finding("critical", "security", resource_id="/sub/rg/r1"),
            self._make_finding("medium", "operational_excellence", resource_id="/sub/rg/r2"),
            self._make_finding("low", "cost_optimization", resource_id="/sub/rg/r3"),
        ]
        phases = build_executive_remediation_roadmap(findings)
        assert len(phases) == 3
        names = [p["name"] for p in phases]
        assert "Phase 1 — Immediate" in names
        assert "Phase 2 — Near Term" in names
        assert "Phase 3 — Strategic" in names

    def test_phase_risk_reduction_values(self) -> None:
        findings = [
            self._make_finding("critical", "security", resource_id="/sub/rg/r1"),
            self._make_finding("medium", "reliability", resource_id="/sub/rg/r2"),
            self._make_finding("low", "cost_optimization", resource_id="/sub/rg/r3"),
        ]
        phases = build_executive_remediation_roadmap(findings)
        phase_map = {p["name"]: p for p in phases}
        assert phase_map["Phase 1 — Immediate"]["risk_reduction"] == "45%"
        assert phase_map["Phase 2 — Near Term"]["risk_reduction"] == "25%"
        assert phase_map["Phase 3 — Strategic"]["risk_reduction"] == "15%"

    def test_deduplicates_same_rule_across_resources(self) -> None:
        findings = [
            self._make_finding("critical", "security", rule_id="WAF-001", resource_id="/rg/r1"),
            self._make_finding("critical", "security", rule_id="WAF-001", resource_id="/rg/r2"),
            self._make_finding("critical", "security", rule_id="WAF-001", resource_id="/rg/r3"),
        ]
        phases = build_executive_remediation_roadmap(findings)
        assert len(phases) == 1
        assert len(phases[0]["items"]) == 1
        assert phases[0]["items"][0]["resource_count"] == 3

    def test_effort_reflects_resource_count(self) -> None:
        findings = [
            self._make_finding(
                "low", "cost_optimization", rule_id="WAF-001", resource_id=f"/rg/r{i}"
            )
            for i in range(7)
        ]
        phases = build_executive_remediation_roadmap(findings)
        assert phases[0]["items"][0]["effort"] == "High"

    def test_phase1_sorted_by_priority_descending(self) -> None:
        findings = [
            self._make_finding("critical", "reliability", rule_id="WAF-A", resource_id="/rg/r1"),
            self._make_finding("critical", "security", rule_id="WAF-B", resource_id="/rg/r2"),
        ]
        phases = build_executive_remediation_roadmap(findings)
        items = phases[0]["items"]
        priorities = [it["priority"] for it in items]
        assert priorities == sorted(priorities, reverse=True)

    def test_items_have_required_keys(self) -> None:
        findings = [self._make_finding("critical", "security")]
        phases = build_executive_remediation_roadmap(findings)
        item = phases[0]["items"][0]
        for key in (
            "title",
            "severity",
            "pillar",
            "resource_count",
            "priority",
            "effort",
            "recommendation",
        ):
            assert key in item


# ── Remediation Playbook tests ─────────────────────────────────────────────────

from waf_reporting.remediation_playbook import (  # noqa: E402
    PlaybookEntry,
    build_remediation_playbook,
    estimate_fix_time,
    expected_risk_reduction,
)


def _make_playbook_finding(
    rule_id: str = "UNKNOWN-RULE-XYZ",
    severity: str = "medium",
    pillar: str = "security",
    resource_type: str = "Microsoft.Web/sites",
    recommendation: str = "Follow the Azure documentation.",
) -> Finding:
    return Finding(
        id=uuid.uuid4(),
        assessment_id=_ASSESSMENT_ID,
        batch_id=_BATCH_ID,
        tenant_id=_TENANT_ID,
        rule_id=rule_id,
        resource_id="/subscriptions/sub/rg/res",
        resource_type=resource_type,
        status=FindingStatus.OPEN,
        severity=Severity(severity),
        pillar=pillar,
        confidence_score=0.9,
        title="Test finding",
        recommendation=recommendation,
        evidence={},
        evaluation_type="deterministic",
        created_at=datetime.now(UTC),
    )


class TestBuildRemediationPlaybook:
    def test_known_finding_returns_playbook_entry(self) -> None:
        f = _make_playbook_finding(
            rule_id="SEC-CR-001",
            resource_type="Microsoft.ContainerRegistry/registries",
        )
        pb = build_remediation_playbook(f)
        assert isinstance(pb, PlaybookEntry)

    def test_unknown_finding_returns_none(self) -> None:
        f = _make_playbook_finding(rule_id="UNKNOWN-RULE-XYZ")
        pb = build_remediation_playbook(f)
        assert pb is None

    def test_portal_steps_not_empty_for_known_rule(self) -> None:
        f = _make_playbook_finding(
            rule_id="SEC-CR-001",
            resource_type="Microsoft.ContainerRegistry/registries",
        )
        pb = build_remediation_playbook(f)
        assert pb is not None
        assert len(pb.portal_steps) > 0

    def test_cli_generation_contains_az_command(self) -> None:
        f = _make_playbook_finding(
            rule_id="SEC-CR-001",
            resource_type="Microsoft.ContainerRegistry/registries",
        )
        pb = build_remediation_playbook(f)
        assert pb is not None
        assert "az acr" in pb.azure_cli

    def test_powershell_generation_not_empty(self) -> None:
        f = _make_playbook_finding(
            rule_id="SEC-DEF-001",
            resource_type="Microsoft.Security/pricings",
        )
        pb = build_remediation_playbook(f)
        assert pb is not None
        assert len(pb.powershell) > 0

    def test_powershell_contains_az_cmdlet(self) -> None:
        f = _make_playbook_finding(
            rule_id="SEC-DEF-001",
            resource_type="Microsoft.Security/pricings",
        )
        pb = build_remediation_playbook(f)
        assert pb is not None
        assert "Set-AzSecurityPricing" in pb.powershell

    def test_bicep_generation_not_empty(self) -> None:
        f = _make_playbook_finding(
            rule_id="SEC-CR-001",
            resource_type="Microsoft.ContainerRegistry/registries",
        )
        pb = build_remediation_playbook(f)
        assert pb is not None
        assert len(pb.bicep) > 0

    def test_terraform_generation_not_empty(self) -> None:
        f = _make_playbook_finding(
            rule_id="SEC-CR-001",
            resource_type="Microsoft.ContainerRegistry/registries",
        )
        pb = build_remediation_playbook(f)
        assert pb is not None
        assert len(pb.terraform) > 0

    def test_all_five_fields_present_and_non_empty(self) -> None:
        f = _make_playbook_finding(
            rule_id="OPS-DIAG-001",
            resource_type="Microsoft.Compute/virtualMachines",
        )
        pb = build_remediation_playbook(f)
        assert pb is not None
        for field in ("portal_steps", "azure_cli", "powershell", "bicep", "terraform"):
            assert len(getattr(pb, field)) > 0

    def test_change_type_field_is_valid(self) -> None:
        f = _make_playbook_finding(
            rule_id="SEC-NET-004",
            resource_type="Microsoft.Web/sites",
        )
        pb = build_remediation_playbook(f)
        assert pb is not None
        assert pb.change_type in ("simple_config", "policy", "network", "architecture")


class TestEstimateFixTime:
    def test_simple_config_rule_returns_15_minutes(self) -> None:
        f = _make_playbook_finding(rule_id="SEC-CR-001")  # simple_config
        assert estimate_fix_time(f) == "15 minutes"

    def test_policy_rule_returns_30_minutes(self) -> None:
        f = _make_playbook_finding(rule_id="CST-COST-TAG-001")  # policy
        assert estimate_fix_time(f) == "30 minutes"

    def test_network_rule_returns_60_minutes(self) -> None:
        f = _make_playbook_finding(rule_id="SEC-NET-004")  # network
        assert estimate_fix_time(f) == "60 minutes"

    def test_architecture_rule_returns_2_4_hours(self) -> None:
        f = _make_playbook_finding(rule_id="OPS-SLOT-001")  # architecture
        assert estimate_fix_time(f) == "2–4 hours"

    def test_unknown_rule_critical_severity_fallback(self) -> None:
        f = _make_playbook_finding(rule_id="UNKNOWN-RULE", severity="critical")
        assert estimate_fix_time(f) == "2–4 hours"

    def test_unknown_rule_medium_severity_fallback(self) -> None:
        f = _make_playbook_finding(rule_id="UNKNOWN-RULE", severity="medium")
        assert estimate_fix_time(f) == "30 minutes"

    def test_unknown_rule_low_severity_fallback(self) -> None:
        f = _make_playbook_finding(rule_id="UNKNOWN-RULE", severity="low")
        assert estimate_fix_time(f) == "15 minutes"


class TestExpectedRiskReduction:
    def test_critical_returns_high(self) -> None:
        f = _make_playbook_finding(severity="critical")
        assert expected_risk_reduction(f) == "High"

    def test_high_returns_medium(self) -> None:
        f = _make_playbook_finding(severity="high")
        assert expected_risk_reduction(f) == "Medium"

    def test_medium_returns_medium(self) -> None:
        f = _make_playbook_finding(severity="medium")
        assert expected_risk_reduction(f) == "Medium"

    def test_low_returns_low(self) -> None:
        f = _make_playbook_finding(severity="low")
        assert expected_risk_reduction(f) == "Low"

    def test_informational_returns_low(self) -> None:
        f = _make_playbook_finding(severity="informational")
        assert expected_risk_reduction(f) == "Low"


# ── Business Impact Analysis ───────────────────────────────────────────────────

from waf_reporting.business_impact_analysis import (  # noqa: E402
    BusinessImpact,
    aggregate_risk_category_levels,
    build_business_impact_analysis,
    calculate_business_impact_score,
)


def _make_impact_finding(
    pillar: str = "security",
    severity: str = "medium",
    rule_id: str = "WAF-TEST-001",
) -> Finding:
    return Finding(
        id=uuid.uuid4(),
        assessment_id=_ASSESSMENT_ID,
        batch_id=_BATCH_ID,
        tenant_id=_TENANT_ID,
        rule_id=rule_id,
        resource_id="/subscriptions/sub/rg/res",
        resource_type="Microsoft.Storage/storageAccounts",
        status=FindingStatus.OPEN,
        severity=Severity(severity),
        pillar=pillar,
        confidence_score=0.9,
        title="Storage account should require secure transfer",
        recommendation="Enable secure transfer requirement.",
        evidence={},
        evaluation_type="deterministic",
        created_at=datetime.now(UTC),
    )


class TestBuildBusinessImpactAnalysis:
    def test_security_finding_returns_security_risk_category(self) -> None:
        f = _make_impact_finding(pillar="security", severity="critical")
        biz = build_business_impact_analysis(f)
        assert biz.risk_category == "Security Risk"

    def test_cost_finding_returns_financial_risk_category(self) -> None:
        f = _make_impact_finding(pillar="cost_optimization", severity="medium")
        biz = build_business_impact_analysis(f)
        assert biz.risk_category == "Financial Risk"

    def test_operational_excellence_returns_operational_risk(self) -> None:
        f = _make_impact_finding(pillar="operational_excellence", severity="high")
        biz = build_business_impact_analysis(f)
        assert biz.risk_category == "Operational Risk"

    def test_reliability_returns_operational_risk(self) -> None:
        f = _make_impact_finding(pillar="reliability", severity="high")
        biz = build_business_impact_analysis(f)
        assert biz.risk_category == "Operational Risk"

    def test_finding_impact_is_not_empty(self) -> None:
        f = _make_impact_finding(pillar="security", severity="high")
        biz = build_business_impact_analysis(f)
        assert len(biz.finding_impact) > 0

    def test_finding_impact_uses_hedged_language(self) -> None:
        f = _make_impact_finding(pillar="security", severity="critical")
        biz = build_business_impact_analysis(f)
        hedged = any(word in biz.finding_impact.lower() for word in ["potential", "may", "could"])
        assert hedged, f"Expected hedged language, got: {biz.finding_impact}"

    def test_priority_critical_is_p1(self) -> None:
        f = _make_impact_finding(severity="critical")
        assert build_business_impact_analysis(f).priority == "P1"

    def test_priority_high_is_p2(self) -> None:
        f = _make_impact_finding(severity="high")
        assert build_business_impact_analysis(f).priority == "P2"

    def test_priority_low_is_p4(self) -> None:
        f = _make_impact_finding(severity="low")
        assert build_business_impact_analysis(f).priority == "P4"

    def test_impact_score_critical_is_100(self) -> None:
        f = _make_impact_finding(severity="critical")
        assert build_business_impact_analysis(f).impact_score == 100

    def test_impact_score_high_is_75(self) -> None:
        f = _make_impact_finding(severity="high")
        assert build_business_impact_analysis(f).impact_score == 75

    def test_impact_score_medium_is_50(self) -> None:
        f = _make_impact_finding(severity="medium")
        assert build_business_impact_analysis(f).impact_score == 50

    def test_returns_business_impact_instance(self) -> None:
        f = _make_impact_finding()
        assert isinstance(build_business_impact_analysis(f), BusinessImpact)


class TestCalculateBusinessImpactScore:
    def test_empty_findings_returns_zero(self) -> None:
        assert calculate_business_impact_score([]) == 0.0

    def test_single_critical_finding_returns_100(self) -> None:
        f = _make_impact_finding(severity="critical")
        assert calculate_business_impact_score([f]) == 100.0

    def test_single_high_finding_returns_75(self) -> None:
        f = _make_impact_finding(severity="high")
        assert calculate_business_impact_score([f]) == 75.0

    def test_single_medium_finding_returns_50(self) -> None:
        f = _make_impact_finding(severity="medium")
        assert calculate_business_impact_score([f]) == 50.0

    def test_average_critical_and_medium(self) -> None:
        findings = [
            _make_impact_finding(severity="critical"),  # 100
            _make_impact_finding(severity="medium"),  # 50
        ]
        assert calculate_business_impact_score(findings) == 75.0

    def test_informational_contributes_zero(self) -> None:
        f = _make_impact_finding(severity="informational")
        assert calculate_business_impact_score([f]) == 0.0

    def test_multiple_findings_averages_correctly(self) -> None:
        findings = [
            _make_impact_finding(severity="high"),  # 75
            _make_impact_finding(severity="low"),  # 25
        ]
        assert calculate_business_impact_score(findings) == 50.0


class TestAggregateRiskCategoryLevels:
    def test_security_critical_raises_security_to_high(self) -> None:
        f = _make_impact_finding(pillar="security", severity="critical")
        levels = aggregate_risk_category_levels([f])
        assert levels["Security Risk"] == "High"

    def test_security_critical_raises_compliance_to_high(self) -> None:
        f = _make_impact_finding(pillar="security", severity="critical")
        levels = aggregate_risk_category_levels([f])
        assert levels["Compliance Risk"] == "High"

    def test_cost_high_raises_financial_risk_to_high(self) -> None:
        f = _make_impact_finding(pillar="cost_optimization", severity="high")
        levels = aggregate_risk_category_levels([f])
        assert levels["Financial Risk"] == "High"

    def test_cost_finding_does_not_raise_security_risk(self) -> None:
        f = _make_impact_finding(pillar="cost_optimization", severity="critical")
        levels = aggregate_risk_category_levels([f])
        assert levels["Security Risk"] == "Low"

    def test_empty_findings_all_low(self) -> None:
        levels = aggregate_risk_category_levels([])
        assert all(v == "Low" for v in levels.values())

    def test_reputation_risk_raised_for_critical_security(self) -> None:
        f = _make_impact_finding(pillar="security", severity="critical")
        levels = aggregate_risk_category_levels([f])
        assert levels["Reputation Risk"] == "High"

    def test_reputation_risk_not_raised_for_low_security(self) -> None:
        f = _make_impact_finding(pillar="security", severity="low")
        levels = aggregate_risk_category_levels([f])
        assert levels["Reputation Risk"] == "Low"

    def test_multiple_pillars_aggregate_correctly(self) -> None:
        findings = [
            _make_impact_finding(pillar="security", severity="medium"),
            _make_impact_finding(pillar="reliability", severity="high"),
            _make_impact_finding(pillar="cost_optimization", severity="low"),
        ]
        levels = aggregate_risk_category_levels(findings)
        assert levels["Security Risk"] == "Medium"
        assert levels["Operational Risk"] == "High"
        assert levels["Financial Risk"] == "Low"

    def test_returns_all_five_categories(self) -> None:
        levels = aggregate_risk_category_levels([])
        expected_keys = {
            "Security Risk",
            "Compliance Risk",
            "Operational Risk",
            "Financial Risk",
            "Reputation Risk",
        }
        assert set(levels.keys()) == expected_keys


# ── Executive Insights ────────────────────────────────────────────────────────

from waf_reporting.executive_insights import (  # noqa: E402
    ExecutiveInsights,
    StrategicRecommendations,
    calculate_insight_confidence,
    generate_executive_insights,
)


def _make_insight_finding(
    pillar: str = "security",
    severity: str = "high",
    resource_type: str = "Microsoft.Storage/storageAccounts",
    rule_id: str = "WAF-INS-001",
    title: str = "Storage secure transfer not enabled",
) -> Finding:
    return Finding(
        id=uuid.uuid4(),
        assessment_id=_ASSESSMENT_ID,
        batch_id=_BATCH_ID,
        tenant_id=_TENANT_ID,
        rule_id=rule_id,
        resource_id="/subscriptions/sub/resourceGroups/rg/providers/Microsoft.Storage/storageAccounts/sa1",
        resource_type=resource_type,
        status=FindingStatus.OPEN,
        severity=Severity(severity),
        pillar=pillar,
        confidence_score=0.9,
        title=title,
        recommendation="Enable secure transfer.",
        evidence={},
        evaluation_type="deterministic",
        created_at=datetime.now(UTC),
    )


# ── TestCalculateInsightConfidence ────────────────────────────────────────────


class TestCalculateInsightConfidence:
    def test_zero_total_returns_low(self) -> None:
        assert calculate_insight_confidence(0, 0) == "Low"

    def test_zero_supporting_returns_low(self) -> None:
        assert calculate_insight_confidence(0, 10) == "Low"

    def test_negative_total_returns_low(self) -> None:
        assert calculate_insight_confidence(1, -1) == "Low"

    def test_high_ratio_and_sufficient_count_returns_high(self) -> None:
        # 6 / 10 = 0.6 >= 0.5 and count >= 3
        assert calculate_insight_confidence(6, 10) == "High"

    def test_high_ratio_but_too_few_returns_medium(self) -> None:
        # 2 / 3 = 0.67 >= 0.5 but count < 3 → not High; count >= 2 → Medium
        assert calculate_insight_confidence(2, 3) == "Medium"

    def test_medium_ratio_returns_medium(self) -> None:
        # 3 / 10 = 0.3 >= 0.25 → Medium
        assert calculate_insight_confidence(3, 10) == "Medium"

    def test_two_supporting_out_of_many_returns_medium(self) -> None:
        # count >= 2 → Medium even if ratio is low
        assert calculate_insight_confidence(2, 20) == "Medium"

    def test_one_of_many_returns_low(self) -> None:
        # 1 / 20 = 0.05, count = 1 → Low
        assert calculate_insight_confidence(1, 20) == "Low"


# ── TestGenerateExecutiveInsights ─────────────────────────────────────────────


class TestGenerateExecutiveInsights:
    def test_empty_findings_returns_executive_insights(self) -> None:
        result = generate_executive_insights([])
        assert isinstance(result, ExecutiveInsights)

    def test_empty_findings_overall_confidence_low(self) -> None:
        result = generate_executive_insights([])
        assert result.overall_confidence == "Low"

    def test_empty_findings_has_at_least_one_observation(self) -> None:
        result = generate_executive_insights([])
        assert len(result.observations) >= 1

    def test_single_finding_returns_insights(self) -> None:
        f = _make_insight_finding()
        result = generate_executive_insights([f])
        assert isinstance(result, ExecutiveInsights)

    def test_single_finding_has_observations(self) -> None:
        f = _make_insight_finding()
        result = generate_executive_insights([f])
        assert len(result.observations) >= 1

    def test_multiple_findings_includes_trend_readiness(self) -> None:
        findings = [
            _make_insight_finding(severity="critical"),
            _make_insight_finding(severity="high"),
            _make_insight_finding(severity="medium"),
        ]
        result = generate_executive_insights(findings)
        types = [o.insight_type for o in result.observations]
        assert "Trend Readiness" in types

    def test_multiple_findings_includes_risk_concentration(self) -> None:
        findings = [_make_insight_finding() for _ in range(5)]
        result = generate_executive_insights(findings)
        types = [o.insight_type for o in result.observations]
        assert "Risk Concentration" in types

    def test_multiple_findings_includes_remediation_leverage(self) -> None:
        findings = [
            _make_insight_finding(severity="critical"),
            _make_insight_finding(severity="high"),
        ]
        result = generate_executive_insights(findings)
        types = [o.insight_type for o in result.observations]
        assert "Remediation Leverage" in types

    def test_governance_insight_generated_when_ops_exceeds_security_by_15(self) -> None:
        # Security score will be low (critical finding), OE will be high (no findings)
        # Provide explicit pillar scores to trigger governance insight
        pillar_scores = [
            ("Security", 50, "Needs Improvement"),
            ("Operational Excellence", 90, "Excellent"),
        ]
        findings = [_make_insight_finding(pillar="security", severity="critical")]
        result = generate_executive_insights(findings, pillar_scores=pillar_scores)
        types = [o.insight_type for o in result.observations]
        assert "Governance" in types

    def test_governance_insight_not_generated_when_gap_too_small(self) -> None:
        pillar_scores = [
            ("Security", 75, "Good"),
            ("Operational Excellence", 80, "Good"),
        ]
        findings = [_make_insight_finding(pillar="security", severity="medium")]
        result = generate_executive_insights(findings, pillar_scores=pillar_scores)
        types = [o.insight_type for o in result.observations]
        assert "Governance" not in types

    def test_mixed_pillar_findings_returns_insights(self) -> None:
        findings = [
            _make_insight_finding(pillar="security", severity="critical"),
            _make_insight_finding(pillar="reliability", severity="high"),
            _make_insight_finding(pillar="cost_optimization", severity="medium"),
            _make_insight_finding(pillar="operational_excellence", severity="low"),
        ]
        result = generate_executive_insights(findings)
        assert isinstance(result, ExecutiveInsights)
        assert len(result.observations) >= 2

    def test_all_observations_have_valid_confidence(self) -> None:
        findings = [_make_insight_finding() for _ in range(3)]
        result = generate_executive_insights(findings)
        valid = {"High", "Medium", "Low"}
        for obs in result.observations:
            assert obs.confidence in valid

    def test_all_observations_have_valid_strategic_priority(self) -> None:
        findings = [_make_insight_finding() for _ in range(3)]
        result = generate_executive_insights(findings)
        valid = {"Immediate", "Near-Term", "Long-Term"}
        for obs in result.observations:
            assert obs.strategic_priority in valid

    def test_returns_executive_insights_instance(self) -> None:
        findings = [_make_insight_finding(severity="high") for _ in range(4)]
        result = generate_executive_insights(findings)
        assert isinstance(result, ExecutiveInsights)
        assert isinstance(result.strategic_recommendations, StrategicRecommendations)


# ── TestExecutiveInsightsNarrative ────────────────────────────────────────────


class TestExecutiveInsightsNarrative:
    def test_narrative_is_non_empty_string(self) -> None:
        findings = [_make_insight_finding(severity="high") for _ in range(3)]
        result = generate_executive_insights(findings)
        assert isinstance(result.assessment_narrative, str)
        assert len(result.assessment_narrative) > 0

    def test_narrative_references_finding_count(self) -> None:
        findings = [_make_insight_finding() for _ in range(5)]
        result = generate_executive_insights(findings)
        assert "5" in result.assessment_narrative

    def test_narrative_uses_conservative_language(self) -> None:
        findings = [_make_insight_finding(severity="critical") for _ in range(3)]
        result = generate_executive_insights(findings)
        forbidden = ["guarantees", "will cause", "certain breach", "proves", "confirms"]
        lower_narr = result.assessment_narrative.lower()
        for word in forbidden:
            assert word not in lower_narr

    def test_empty_findings_narrative_is_fallback_string(self) -> None:
        result = generate_executive_insights([])
        assert len(result.assessment_narrative) > 10

    def test_strategic_recommendations_all_populated(self) -> None:
        findings = [_make_insight_finding(severity="critical") for _ in range(2)]
        result = generate_executive_insights(findings)
        recs = result.strategic_recommendations
        assert len(recs.immediate_focus) > 0
        assert len(recs.near_term_focus) > 0
        assert len(recs.long_term_focus) > 0

    def test_near_term_mentions_monitoring_for_ops_pillar(self) -> None:
        findings = [_make_insight_finding(pillar="operational_excellence", severity="medium")]
        result = generate_executive_insights(findings)
        assert "monitor" in result.strategic_recommendations.near_term_focus.lower()

    def test_long_term_mentions_governance_for_cost_pillar(self) -> None:
        findings = [_make_insight_finding(pillar="cost_optimization", severity="low")]
        result = generate_executive_insights(findings)
        assert "governance" in result.strategic_recommendations.long_term_focus.lower()


# ── TestExcelSheetExecutiveInsights ───────────────────────────────────────────


class TestExcelSheetExecutiveInsights:
    def _gen_excel(self, findings: list[Finding]) -> bytes:
        agg = _make_aggregated(total_findings=len(findings))
        return ExcelGenerator().generate(agg, findings)

    def test_ai_executive_insights_sheet_created(self) -> None:
        import openpyxl

        findings = [_make_insight_finding() for _ in range(3)]
        data = self._gen_excel(findings)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "AI Executive Insights" in wb.sheetnames

    def test_sheet_has_expected_headers(self) -> None:
        import openpyxl

        findings = [_make_insight_finding(severity="high") for _ in range(3)]
        data = self._gen_excel(findings)
        wb = openpyxl.load_workbook(io.BytesIO(data))
        ws = wb["AI Executive Insights"]
        flat = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value]
        assert any("Insight Type" in v for v in flat)

    def test_sheet_generates_without_findings(self) -> None:
        import openpyxl

        data = self._gen_excel([])
        wb = openpyxl.load_workbook(io.BytesIO(data))
        assert "AI Executive Insights" in wb.sheetnames


# ── TestPdfSectionExecutiveInsights ──────────────────────────────────────────


class TestPdfSectionExecutiveInsights:
    def _gen_pdf(self, findings: list[Finding]) -> bytes:
        agg = _make_aggregated(total_findings=len(findings))
        return PdfGenerator().generate(agg, findings)

    def test_pdf_generates_with_multiple_findings(self) -> None:
        findings = [_make_insight_finding(severity="critical") for _ in range(3)]
        pdf = self._gen_pdf(findings)
        assert len(pdf) > 1000

    def test_pdf_generates_with_empty_findings(self) -> None:
        pdf = self._gen_pdf([])
        assert len(pdf) > 1000

    def test_pdf_generates_with_mixed_pillars(self) -> None:
        findings = [
            _make_insight_finding(pillar="security", severity="critical"),
            _make_insight_finding(pillar="reliability", severity="high"),
            _make_insight_finding(pillar="operational_excellence", severity="medium"),
            _make_insight_finding(pillar="cost_optimization", severity="low"),
        ]
        pdf = self._gen_pdf(findings)
        assert len(pdf) > 1000
